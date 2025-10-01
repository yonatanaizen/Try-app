import pandas as pd
import numpy as np
from datetime import time
class build_p1():
    def __init__(self, path):
        self.path = path
        data = pd.read_excel(path)
        data = data.iloc[:, 0:17]
        data.columns = data.iloc[0:1].values.tolist()[0]
        self.data = data.iloc[1:]
        # self.data['זמן בפועל יציאת רכבות מחיל האוויר על פי המצלמות']=pd.to_datetime(self.data['זמן בפועל יציאת רכבות מחיל האוויר על פי המצלמות'], errors='coerce').dt.time
        self.recover_data()

        self.filter()  # Automatically filters during init
        self.create_new()
        self.chose_columns()
        self.recover_data_correction()

    def recover_data(self):
        '''
        Function that need to change only if theres a prbolome with the arival of the train
        '''
        data=self.data
        ls_columns = data.columns.tolist()



        data.loc[(~(data[ls_columns[1]].isna()))&(data[ls_columns[2]].isna()), ls_columns[2]] = data[~(data[ls_columns[1]].isna())][ls_columns[1]]

        data.loc[(~(data[ls_columns[2]].isna()))&(data[ls_columns[6]].isna()), ls_columns[6]] = data[(~(data[ls_columns[2]].isna()))&(data[ls_columns[6]].isna())][ls_columns[2]]


        # data.loc[(~(data[ls_columns[1]].isna()))&(data[ls_columns[6]].isna()), ls_columns[6]] = data[(~(data[ls_columns[1]].isna()))&(data[ls_columns[6]].isna())][ls_columns[1]]

        data.loc[(~(data[ls_columns[1]].isna()))&(data[ls_columns[11]].isna()), ls_columns[11]] = pd.to_datetime('0:00').time()


        self.data=data

    def filter(self):
        self.data_r = self.data[~self.data['קרון משני'].isna()].reset_index(drop=True)

    def time_y(self, a):
        if str(a) == 'nan':
            return a
        else:
            if len(str(int(round((a - int(a)) * 60, 4)))) == 1 and round((a - int(a)) * 60, 4) != 0:
                res = '00.' + str(int(a)) + '.0' + str(int(round(a - int(a), 4) * 60))
            elif len(str(int(round((a - int(a)) * 60, 4)))) == 1 and round((a - int(a)) * 60, 4) == 0:
                res = '00.' + str(int(a)) + '.00'
            else:
                res = '00.' + str(int(a)) + '.' + str(int(round(a - int(a), 4) * 60))
                # res=a if  str(a)=='nan' else  str(int(a))+'.'+str(int(round(a-int(a),4)*60))
            return res
    def find_loc(self,N):
        'find the location'
        fr = list(self.data_r.columns)[N]
        a = 'None'
        if fr.find('חיל האוויר') != -1: a = 'חיל האוויר'
        if fr.find('הר הרצל') != -1: a = 'הר הרצל'
        if fr.find('נווה יעקב') != -1: a = 'נווה יעקב'
        if fr.find('הדסה עין כרם') != -1: a = 'הדסה עין כרם'
        if fr.find ('גבעת המיבתר')!= -1 : a='גבעת המיבתר'
        if fr.find('גבעת המבתר') != -1: a = 'גבעת המבתר'
        if fr.find('תחנה המרכזית') != -1: a = 'התחנה המרכזית'
        if fr.find('התחנה המרזית') != -1: a = 'התחנה המרזית'
        return a

    def create_new(self):
        df_res = pd.DataFrame()
        df_res['lead'] = self.data_r['קרון מוביל']
        df_res['sec'] = self.data_r['קרון משני']
        # print(self.path.name)
        # df_res['date'] = self.path.split('\\')[-1][:8]
        df_res['date'] = self.path.name[:8]
        # print(df_res['date'].iloc[0])

        df_res['day'] = (pd.to_datetime(df_res['date'].iloc[0]).weekday() + 1) % 6


        df_res['cluster'] = 3 if df_res['day'].iloc[0] < 6 else 1
        # df_res['day'] = self.path.split('\\')[-1][:2]

        df_res['name'] = df_res[['lead', 'sec']].max(axis=1).astype(int).astype(str) + '-' + df_res[
            ['lead', 'sec']].min(axis=1).astype(int).astype(str)
        try:
            df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מחיל האוויר על פי המצלמות']
        except:
            try:
                df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מהתחנה המרכזית על פי המצלמות']
            except:
                df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מנווה יעקב על פי המצלמות']

        # print(df_res['exit'])
        try:
            try:
                df_res['arrivel'] = self.data_r['שעת הגעה להר הרצל']
            except:
                df_res['arrivel'] = self.data_r['שעת הגעה להתחנה המרכזית']
        except:
            try:
                try:
                    try:
                        df_res['arrivel'] = self.data_r['שעת הגעה להדסה']
                    except:
                        df_res['arrivel'] = self.data_r['שעת הגעה להדסה  עין כרם']

                except:
                    df_res['arrivel'] = self.data_r['שעת הגעה להדסה עין כרם']
            except:
                try:
                    df_res['arrivel'] = self.data_r['שעת הגעה לגבעת המיבתר']
                except:
                    df_res['arrivel'] = self.data_r['שעת הגעה לגבעת המבתר']

        df_res.loc[df_res['arrivel'].astype(str) == '1 day, 0:00:00', 'arrivel'] = '00:00:00'
        df_res.loc[df_res['exit'].astype(str) == '1 day, 0:00:00', 'exit'] = '00:00:00'

        df_res['time'] = self.data_r['זמן נסיעה']
        # print(self.data_r.columns)


        # df_res['from'] = 'נווה יעקב צפון'
        df_res['from'] = self.find_loc(2)
        # df_res['to'] = 'הדסה עין כרם'
        df_res['to'] = self.find_loc(6)

        df_res['direction'] = self.find_loc(6)
        df_res['comments'] = np.nan
        df_res['hours'] = np.nan
        df_res['exit'] = pd.to_datetime(df_res['exit'], format='%H:%M:%S', errors='coerce').dt.time
        df_res.loc[(df_res['exit'] < time(6, 0, 0)) & (df_res['exit'] >= time(5, 30, 0)), 'hours'] = '05:30-06:00'
        df_res.loc[(df_res['exit'] < time(7, 0, 0)) & (df_res['exit'] >= time(6, 0, 0)), 'hours'] = '06:00-07:00'
        df_res.loc[(df_res['exit'] < time(19, 0, 0)) & (df_res['exit'] >= time(7, 0, 0)), 'hours'] = '07:00-19:00'
        df_res.loc[(df_res['exit'] < time(20, 0, 0)) & (df_res['exit'] >= time(19, 0, 0)), 'hours'] = '19:00-20:00'
        df_res.loc[(df_res['exit'] < time(22, 0, 0)) & (df_res['exit'] >= time(20, 0, 0)), 'hours'] = '20:00-22:00'
        df_res.loc[(df_res['exit'] <= time(23, 59, 0)) & (df_res['exit'] >= time(22, 0, 0)), 'hours'] = '22:00-24:00'

        df_res['Freq'] = pd.to_timedelta(df_res['exit'].astype(str)) - pd.to_timedelta(
            df_res['exit'].astype(str).shift(1))

        df_res['Freq_sec'] = df_res['Freq'].dt.total_seconds()
        df_res.loc[df_res['Freq_sec'] < 0, 'Freq_sec'] = df_res[df_res['Freq_sec'] < 0]['Freq_sec'] + 24 * 60 * 60
        df_res['Time'] = df_res['Freq_sec'] / 60
        time_y = lambda a: a if str(a) == 'nan' else str(int(a)) + '.' + str(int(round(a - int(a), 4) * 60))

        df_res['TimeC'] = df_res['Time'].apply(time_y)
        # df_res['TimeC']=df_res['Time'].apply(self.time_y)
        df_res['time']=pd.to_datetime(df_res['time'], format='%H:%M:%S', errors='coerce').dt.time
        # print( pd.to_datetime(df_res['time'], format='%H:%M:%S', errors='coerce').dt.time)

        df_res['distrbution'] = 'מעל 60'
        df_res.loc[(df_res['time'] >= time(0, 55)) & (df_res['time'] < time(1)), 'distrbution'] = '55-60'
        df_res.loc[(df_res['time'] >= time(0, 50)) & (df_res['time'] < time(0, 55)), 'distrbution'] = '50-55'
        df_res.loc[(df_res['time'] >= time(0, 45)) & (df_res['time'] < time(0, 50)), 'distrbution'] = '45-50'
        df_res.loc[(df_res['time'] >= time(0, 40)) & (df_res['time'] < time(0, 45)), 'distrbution'] = '40-45'
        df_res.loc[df_res['time'].isna(), 'distrbution'] = np.nan

        df_res['travel'] = 'נסיעה מלאה'
        df_res.loc[(df_res['exit'].isna()) | (df_res['arrivel'].isna()), 'travel'] = 'נסיעה חלקית'
        self.result = df_res

    def chose_columns(self):
        self.end = self.result[
            ['date', 'day', 'cluster', 'lead', 'sec', 'name', 'exit', 'arrivel', 'time', 'from', 'to', 'direction',
             'comments', 'hours', 'TimeC', 'distrbution', 'travel']]

    def  recover_data_correction(self):
        data=self.end
        # data.loc[(data['time'] == pd.to_datetime('0:00').time()) & (data['exit'] != data['arrivel']), 'exit'] = np.nan
        #
        # data.loc[(data['time']==pd.to_datetime('0:00').time())&(data['exit']==data['arrivel']),'arrivel']=np.nan

        data.loc[(data['time'] == pd.to_datetime('0:00').time()) & (pd.to_datetime(data['exit'].astype(str)).dt.time != pd.to_datetime(data['arrivel'].astype(str)).dt.time), 'exit'] = np.nan

        data.loc[(data['time'] == pd.to_datetime('0:00').time()) & (pd.to_datetime(data['exit'].astype(str)).dt.time == pd.to_datetime(data['arrivel'].astype(str)).dt.time), 'arrivel'] = np.nan

        data.loc[(data['time'] == pd.to_datetime('0:00').time()) , 'time'] = np.nan

        self.end=data

    # def recover_data_correction(self):
    #     '''
    #     Function that corect the recover and return the time that wasnt apper to nan theres 2 option that not arive that not depart
    #     :return:
    #     '''
    #     data = self.data
    #     'Option a'
    #     print(data['time'])
    #     # data.loc[data['time']]
    #
    #     # print(9)


class build_p2():
    def __init__(self, path):
        self.path = path
        data = pd.read_excel(path)
        data = data.iloc[:, 18:35]
        data.columns = data.iloc[0:1].values.tolist()[0]
        self.data = data.iloc[1:]
        self.recover_data()
        self.filter()  # Automatically filters during init
        self.create_new()
        self.chose_columns()
        self.recover_data_correction()

    def recover_data(self):
        '''
        Function that need to change only if theres a prbolome with the arival of the train
        '''
        data = self.data
        ls_columns = data.columns.tolist()


        data.loc[(~(data[ls_columns[1]].isna()))&(data[ls_columns[2]].isna()), ls_columns[2]] = data[~(data[ls_columns[1]].isna())][ls_columns[1]]

        data.loc[(~(data[ls_columns[2]].isna()))&(data[ls_columns[6]].isna()), ls_columns[6]] = data[(~(data[ls_columns[2]].isna()))&(data[ls_columns[6]].isna())][ls_columns[2]]


        # data.loc[(~(data[ls_columns[1]].isna()))&(data[ls_columns[6]].isna()), ls_columns[6]] = data[(~(data[ls_columns[1]].isna()))&(data[ls_columns[6]].isna())][ls_columns[1]]

        data.loc[(~(data[ls_columns[1]].isna()))&(data[ls_columns[11]].isna()), ls_columns[11]] = pd.to_datetime('0:00').time()

        self.data = data

    def filter(self):
        self.data_r = self.data[~self.data['קרון משני'].isna()].reset_index(drop=True)

    def time_y(self, a):
        if str(a) == 'nan':
            return a
        else:
            if len(str(int(round((a - int(a)) * 60, 4)))) == 1 and round((a - int(a)) * 60, 4) != 0:
                res = '00.' + str(int(a)) + '.0' + str(int(round(a - int(a), 4) * 60))
            elif len(str(int(round((a - int(a)) * 60, 4)))) == 1 and round((a - int(a)) * 60, 4) == 0:
                res = '00.' + str(int(a)) + '.00'
            else:
                res = '00.' + str(int(a)) + '.' + str(int(round(a - int(a), 4) * 60))
                # res=a if  str(a)=='nan' else  str(int(a))+'.'+str(int(round(a-int(a),4)*60))
            return res

    def find_loc(self,N):
        'find the location'
        fr = list(self.data_r.columns)[N]
        a = 'None'
        if fr.find('חיל האוויר') != -1: a = 'חיל האוויר'
        if fr.find('הר הרצל') != -1: a = 'הר הרצל'
        if fr.find('נווה יעקב') != -1: a = 'נווה יעקב'
        if fr.find('הדסה עין כרם') != -1: a = 'הדסה עין כרם'
        if fr.find ('גבעת המיבתר')!= -1 : a='גבעת המיבתר'
        if fr.find('גבעת המבתר') != -1: a = 'גבעת המבתר'
        if fr.find('הדסה') != -1: a = 'הדסה'
        if fr.find('תחנה המרכזית') != -1: a = 'התחנה המרכזית'
        if fr.find('התחנה המרזית') != -1: a = 'התחנה המרזית'

        return a

    def create_new(self):
        df_res = pd.DataFrame()
        df_res['lead'] = self.data_r['קרון מוביל']
        df_res['sec'] = self.data_r['קרון משני']

        # df_res['date'] = self.path.split('\\')[-1][:8]
        df_res['date'] = self.path.name[:8]
        df_res['day'] = (pd.to_datetime(df_res['date'].iloc[0]).weekday() + 1) % 6
        # df_res['day'] = str(df_res['date'].iloc[0])[:2]

        df_res['cluster'] = 3 if df_res['day'].iloc[0] < 6 else 1
        # df_res['day'] = self.path.split('\\')[-1][:2]

        df_res['name'] = df_res[['lead', 'sec']].max(axis=1).astype(int).astype(str) + '-' + df_res[
            ['lead', 'sec']].min(axis=1).astype(int).astype(str)
        try:
            try:
                try:
                    df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מהר הרצל על פי המצלמות']
                except:
                    df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מהתחנה המרכזית על פי המצלמות']

            except:
                df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מגבעת המבתר על פי המצלמות']
        except:
            try:
                try:
                    df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מהדסה על פי המצלמות']
                except:
                    df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מהדסה עין כרם על פי המצלמות']


            except:
                df_res['exit'] = self.data_r['זמן בפועל יציאת רכבות מגבעת המביתר על פי המצלמות']


        try:
            df_res['arrivel'] = self.data_r['שעת הגעה לחיל האוויר']
        except:
            try:
                try:
                    try:
                        df_res['arrivel'] = self.data_r['שעת הגעה לתחנה מרכזית']
                    except:
                        df_res['arrivel'] = self.data_r['שעת הגעה לנווה יעקב']
                except:
                    try:
                        df_res['arrivel'] = self.data_r['שעת הגעה להתחנה מרכזית']
                    except:
                        df_res['arrivel'] = self.data_r['שעת הגעה להתחנה המרכזית']
            except:
                df_res['arrivel'] = self.data_r['שעת הגעה להתחנה המרזית']

        df_res.loc[df_res['arrivel'].astype(str) == '1 day, 0:00:00','arrivel'] = '00:00:00'
        df_res.loc[df_res['exit'].astype(str) == '1 day, 0:00:00', 'exit'] = '00:00:00'

        # df_res.loc[df_res['arrivel']=='1/1/1900  12:00:00 AM','arrivel']='00:00:00'
        # df_res.loc[df_res['exit'] == '1/1/1900  12:00:00 AM', 'exit'] = '00:00:00'
        df_res['time'] = self.data_r['זמן נסיעה']
        # df_res['from'] = 'הדסה עין כרם'
        # df_res['to'] = 'נווה יעקב צפון'
        # df_res['direction'] = 'נווה יעקב צפון'
        # df_res['from'] = 'נווה יעקב צפון'
        df_res['from'] = self.find_loc(2)
        # df_res['to'] = 'הדסה עין כרם'
        df_res['to'] = self.find_loc(6)

        df_res['direction'] = self.find_loc(6)
        df_res['comments'] = np.nan
        df_res['hours'] = np.nan
        df_res['exit'] = pd.to_datetime(df_res['exit'], format='%H:%M:%S', errors='coerce').dt.time
        df_res.loc[(df_res['exit'] < time(6, 0, 0)) & (df_res['exit'] >= time(5, 30, 0)), 'hours'] = '05:30-06:00'
        df_res.loc[(df_res['exit'] < time(7, 0, 0)) & (df_res['exit'] >= time(6, 0, 0)), 'hours'] = '06:00-07:00'
        df_res.loc[(df_res['exit'] < time(19, 0, 0)) & (df_res['exit'] >= time(7, 0, 0)), 'hours'] = '07:00-19:00'
        df_res.loc[(df_res['exit'] < time(20, 0, 0)) & (df_res['exit'] >= time(19, 0, 0)), 'hours'] = '19:00-20:00'
        df_res.loc[(df_res['exit'] < time(22, 0, 0)) & (df_res['exit'] >= time(20, 0, 0)), 'hours'] = '20:00-22:00'
        df_res.loc[(df_res['exit'] <= time(23, 59, 0)) & (df_res['exit'] >= time(22, 0, 0)), 'hours'] = '22:00-24:00'

        df_res['Freq'] = pd.to_timedelta(df_res['exit'].astype(str)) - pd.to_timedelta(
            df_res['exit'].astype(str).shift(1))

        df_res['Freq_sec'] = df_res['Freq'].dt.total_seconds()
        df_res.loc[df_res['Freq_sec'] < 0, 'Freq_sec'] = df_res[df_res['Freq_sec'] < 0]['Freq_sec'] + 24 * 60 * 60
        df_res['Time'] = df_res['Freq_sec'] / 60
        time_y = lambda a: a if str(a) == 'nan' else str(int(a)) + '.' + str(int(round(a - int(a), 4) * 60))

        df_res['TimeC'] = df_res['Time'].apply(time_y)
        df_res['time'] = pd.to_datetime(df_res['time'], format='%H:%M:%S', errors='coerce').dt.time

        # df_res['TimeC']=df_res['Time'].apply(self.time_y)
        df_res['distrbution'] = 'מעל 60'
        df_res.loc[(df_res['time'] >= time(0, 55)) & (df_res['time'] < time(1)), 'distrbution'] = '55-60'
        df_res.loc[(df_res['time'] >= time(0, 50)) & (df_res['time'] < time(0, 55)), 'distrbution'] = '50-55'
        df_res.loc[(df_res['time'] >= time(0, 45)) & (df_res['time'] < time(0, 50)), 'distrbution'] = '45-50'
        df_res.loc[(df_res['time'] >= time(0, 40)) & (df_res['time'] < time(0, 45)), 'distrbution'] = '40-45'
        df_res.loc[df_res['time'].isna(), 'distrbution'] = np.nan

        df_res['travel'] = 'נסיעה מלאה'
        df_res.loc[(df_res['exit'].isna()) | (df_res['arrivel'].isna()), 'travel'] = 'נסיעה חלקית'
        self.result = df_res

    def chose_columns(self):
        self.end = self.result[
            ['date', 'day', 'cluster', 'lead', 'sec', 'name', 'exit', 'arrivel', 'time', 'from', 'to', 'direction',
             'comments', 'hours', 'TimeC', 'distrbution', 'travel']]

        # print(9)

    def recover_data_correction(self):

        data = self.end
        data.loc[(data['time'] == pd.to_datetime('0:00').time()) & (pd.to_datetime(data['exit'].astype(str)).dt.time != pd.to_datetime(data['arrivel'].astype(str)).dt.time), 'exit'] = np.nan

        data.loc[(data['time'] == pd.to_datetime('0:00').time()) & (pd.to_datetime(data['exit'].astype(str)).dt.time == pd.to_datetime(data['arrivel'].astype(str)).dt.time), 'arrivel'] = np.nan
        data.loc[(data['time'] == pd.to_datetime('0:00').time()), 'time'] = np.nan

        self.end = data


def all_t(path):

    a=build_p1(path=path)
    b=build_p2(path=path)


    r=pd.DataFrame(['x'],columns=['date'],index=[len(a.end)])
    a.end=a.end._append(r)

    b.end=b.end._append(r)
    b.end=b.end._append(r)


    end=pd.concat([a.end,b.end])
    end['day']=end['date'].astype(str).apply(lambda a:a[:2])



    def rep(a):
        try:
            if str(a) == 'nan':
                return a
            else:
                r = str(a).split('.')
                return time(hour=0, minute=int(r[0]), second=int(r[1]))

        except:
            print(a)

    def find_the_gap(a):
        'function that solve the prblome that  time apper with year in part of the time '
        try:
            return  str(a).split()[1]
        except:
            return a

        # return '0'+ str(a.split()[1]) if len(a.split()) > 1 else a




        # try:
        #     pd.to_datetime(a)
        # except:
        #     print(a)


    end['date'] = end['date'].str.replace('.', '/')
    print( end['TimeC'].sample(6))
    # print(end['TimeC'].max())


    end['TimeC'] = end['TimeC'].apply(rep)
    end['TimeC'] = end['TimeC'].shift(-1)


    end=end.reset_index(drop=True)
    X_list=end[end['date']=='x'].index.tolist()

    end.columns = ['תאריך','יום','סיווג',"מס' קרון מוביל","מס'קרון משני",'שם רכבת','שעת יציאה','שעת הגעה','משך נסיעה','מוצא','יעד','כיוון','הערות','טווח שעות','תדירות','התפלגות משך נסיעה','הגדרות נסיעה']

    # end['שעת הגעה']=pd.to_datetime(end['שעת הגעה'],errors='coerce',format='%Y-%m-%d %H:%M:%S').dt.time
    # pd.to_datetime(X, ).dt.time
    end['שעת יציאה']=end['שעת יציאה'].apply(find_the_gap)
    end['שעת הגעה']=end['שעת הגעה'].apply(find_the_gap)
    # end['שעת יציאה'] = pd.to_datetime(end['שעת יציאה'],errors='coerce',format='%Y-%m-%d %H:%M:%S').dt.time
    # print(end['שעת יציאה'])
    # print(end[['תאריך']].head())

    con=lambda a: a[:-2] +'20'+a[-2:]  if len(a)>3 else a # solve the problome there is only 2 digit for year
    end['תאריך']=end['תאריך'].apply(con)
    # end['תאריך'] = pd.to_datetime(end['תאריך'], format='%d/%m/%y').dt.date

    import openpyxl

    filename = 'styled_output.xlsx'
    end.to_excel(filename, index=False)
    # openpyxl.load_workbook(end)

    # Load workbook to apply styling
    wb = openpyxl.load_workbook(filename)

    from openpyxl.styles import Font
    from openpyxl.styles import Font, PatternFill

    sheet_obj = wb.active

    red_bold_font = Font(color="FF0000", bold=True)
    blue_fill = PatternFill(start_color="B2B2E0", end_color="B2B2E0", fill_type="solid")
    grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    # Apply styles to "משך נסיעה" column (assumed column B)

    for row in sheet_obj.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.font = red_bold_font
            cell.fill = blue_fill
    for i in range(1, 17):
        # sheet_obj.cell( row=1,column = i).font= Font(color="FF0000", bold=True)
        sheet_obj.cell(row=1, column=i).fill = blue_fill

    for i in X_list:
        for j in range(1, 18):
            sheet_obj.cell(row=i + 2, column=i).font = Font(color='000000', bold=True)

            sheet_obj.cell(row=i + 2, column=j).fill = grey_fill

    sheet_obj.sheet_view.rightToLeft = True

    # for col in sheet_obj.iter_columns(min_row=2, min_col=1, max_col=14):
    #     for cell in col:
    #         # Apply the styles to each cell
    #         cell.font = red_bold_font
    #         cell.fill = blue_fill
    # cell.font = Font(color="FF0000", bold=True)  # Red + Bold
    return  wb
    # wb.save(r"C:\Users\yonatana\Downloads\try_new.xlsx")
