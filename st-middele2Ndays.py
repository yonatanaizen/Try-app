import streamlit as st
import pandas as pd
import io
from process_the_data import   all_t
import openpyxl

import copy

def append_sheet_with_styles(source_ws, target_ws, start_row):
    row_offset = start_row - 1

    for row in source_ws.iter_rows():
        for cell in row:
            new_cell = target_ws.cell(
                row=cell.row + row_offset,
                column=cell.col_idx,
                value=cell.value
            )
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = copy.copy(cell.number_format)
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)

    # Copy column widths
    for col_letter, dim in source_ws.column_dimensions.items():
        target_ws.column_dimensions[col_letter].width = dim.width

    # Copy merged cells
    for merged_range in source_ws.merged_cells.ranges:
        target_ws.merge_cells(
            start_row=merged_range.min_row + row_offset,
            start_column=merged_range.min_col,
            end_row=merged_range.max_row + row_offset,
            end_column=merged_range.max_col
        )

    return target_ws.max_row + 2  # Return new starting row


def data_process(uploaded_file:pd.DataFrame):
    '''
    funcion that got data and do the process
    :param data: dataframe
    :return: df new column 9
    '''
    if uploaded_file.name.endswith('.csv'):
        data=pd.read_csv(uploaded_file)
    else:
        data=pd.read_excel(uploaded_file)

    data['new_c']=9
    return data
from datetime import datetime

def extract_date(file):
    try:
        # Assumes filename is like "01.02.2025.xlsx"
        date_part = file.name.rsplit('.', 2)[0]  # "01.02.2025"
        return date_part
    except Exception as e:
        st.warning(f"Filename format invalid: {file.name}")
        return datetime.min  # So invalid files go to the start

st.set_page_config(page_title="Montly result", layout="centered")
st.title('Monthly result')
st.write("Upload your CSV or XLSX file")

uploaded_files = st.file_uploader("Choose a file", type=["csv", "xlsx"],accept_multiple_files=True)
uploaded_files = sorted(uploaded_files, key=extract_date)

# This entire block will ONLY run if a file has been uploaded
if uploaded_files:
    merged_wb = openpyxl.Workbook()
    merged_ws = merged_wb.active
    merged_ws.title = "Merged Data"

    first = True  # To remove the default empty sheet later

    c=0
    for uploaded_file in uploaded_files:
        try:
            wb = all_t(uploaded_file)
        except:
            st.write(f"There is a problem with the file {uploaded_file.name}:")
            break
        for sheet in wb.worksheets:
            c2 = 0

            for row in sheet.iter_rows(values_only=True):
                if c2==0 and c>0:
                    pass
                else:
                    merged_ws.append(row)
                c2=c2+1
        c=c+1
                # current_row = append_sheet_with_styles(sheet, merged_ws, current_row)

        # Optional: add an empty row between files
        # merged_ws.append([None] * merged_ws.max_column)

    # Clean up: remove the default empty row if it was never used
    if first and merged_ws.max_row == 1 and all(cell.value is None for cell in merged_ws[1]):
        merged_ws.delete_rows(1)
    first = False
    from openpyxl.styles import Font, PatternFill


    red_bold_font = Font(color="FF0000", bold=True)
    blue_fill = PatternFill(start_color="B2B2E0", end_color="B2B2E0", fill_type="solid")
    grey_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")

    # Apply styles to "משך נסיעה" column (assumed column B)

    for row in merged_ws.iter_rows(min_row=2, min_col=9, max_col=9):
        for cell in row:
            cell.font = red_bold_font
            cell.fill = blue_fill
    for i in range(1, 17):
        # sheet_obj.cell( row=1,column = i).font= Font(color="FF0000", bold=True)
        merged_ws.cell(row=1, column=i).fill = blue_fill
    l=[i for i in range(1,merged_ws.max_row+1) if merged_ws.cell(row=i, column=1).value == 'x'  ]

    for i in l:
        for j in range(1, 18):
            merged_ws.cell(row=i , column=i).font = Font(color='000000', bold=True)

            merged_ws.cell(row=i , column=j).fill = grey_fill
    merged_ws.sheet_view.rightToLeft = True

    # Download as single Excel file
    excel_buffer = io.BytesIO()
    merged_wb.save(excel_buffer)
    excel_buffer.seek(0)

    st.markdown("---")
    st.write("### Download the Merged File:")

    st.download_button(
        label="Download Merged XLSX",
        data=excel_buffer,
        file_name="merged_data.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_merged_excel"
    )
